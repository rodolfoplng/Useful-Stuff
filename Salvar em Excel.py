df.to_excel('File_Name.xlsx', index=False)

files_names = ['File_Name.xlsx']

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Define border style
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))


def format_excel_file(file_path):
    # Load Excel file
    wb = load_workbook(file_path)
    
    for ws in wb.worksheets:
        # Define header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # Apply formatting to header
        for col_num, column_title in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        # Apply borders to data cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

        # Adjust column widths
        for col_num, column_cells in enumerate(ws.iter_cols(min_row=1, max_col=ws.max_column), 1):
            max_length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2
    
    # Save formatted Excel file
    wb.save(file_path)

# Apply formatting to each file in the list
for file_path in files_names:

    format_excel_file(file_path)
